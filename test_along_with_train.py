import ast
import os
import numpy as np
from openpyxl import Workbook, load_workbook
import tensorflow as tf

from param import args
from env.env import Environment

class TestManager:
    def __init__(self, filename):
        self.filename = filename
        self._init_excel()
        self.wb = load_workbook(self.filename)
        self.ws = self.wb["results"]

    def _init_excel(self):
        """创建 Excel 文件及路径"""
        folder = os.path.dirname(self.filename)
        if folder != "" and not os.path.exists(folder):
            os.makedirs(folder)

        if not os.path.exists(self.filename):
            wb = Workbook()
            ws = wb.active
            ws.title = "results"
            ws.append(['w1', 'w2', 'w3', "avg_total", "avg_d", "avg_h", "avg_s"])
            wb.save(self.filename)

    def save(self, w1, w2, w3, avg_total, avg_d, avg_h, avg_s):
        """自动追加一行测试结果"""
        self.ws.append([w1, w2, w3, avg_total, avg_d, avg_h, avg_s])
        self.wb.save(self.filename)

    def run_test(self, agent1, agent2, agent3, epoch, filename=f'./env/uniform_weights_500.txt'):
        """
        运行一次完整测试流程，并写入 Excel
        ----------------------------------------------------
        env: 环境对象
        agent: 训练好的智能体
        epoch: 当前训练的 epoch
        num_exp: 测试重复次数
        """
        totals, ds, hs, ss = [], [], [], []
        i = 0
        w1 = w2 = w3 = 0.0
        tf.set_random_seed(args.seed)
        env = Environment()
        with open(filename, 'r', encoding='utf-8') as f:
            content = f.read()
            # 使用 ast.literal_eval 安全地将字符串解析为 Python 列表
            weights = ast.literal_eval(content)

        for w in weights:
            print('xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx')
            env.seed(42 + i)
            env.reset(max_time=100)
            done = False
            print(w)

            while not done:
                obs = env.observe()
                job_dags, frontier_nodes, free_mecs, \
                    node_action_map, mec_action_map = obs
                if len(frontier_nodes) == 0 or len(free_mecs) == 0:
                    return None, None
                node_act_probs1, job_act_probs1, \
                    node_inputs, job_inputs, \
                    node_valid_mask, job_valid_mask, \
                    gcn_mats, gcn_masks, summ_mats, \
                    running_dags_mat, dag_summ_backward_map, \
                    job_dags_changed = \
                    agent1.invoke_model(obs)
                node_act_probs2, job_act_probs2, \
                    node_inputs, job_inputs, \
                    node_valid_mask, job_valid_mask, \
                    gcn_mats, gcn_masks, summ_mats, \
                    running_dags_mat, dag_summ_backward_map, \
                    job_dags_changed = \
                    agent2.invoke_model(obs)
                node_act_probs3, job_act_probs3, \
                    node_inputs, job_inputs, \
                    node_valid_mask, job_valid_mask, \
                    gcn_mats, gcn_masks, summ_mats, \
                    running_dags_mat, dag_summ_backward_map, \
                    job_dags_changed = \
                    agent3.invoke_model(obs)
                w1 = w[0]
                w2 = w[1]
                w3 = w[2]
                eps = 1e-13
                node_logits1 = np.log(node_act_probs1 + eps)
                node_logits2 = np.log(node_act_probs2 + eps)
                node_logits3 = np.log(node_act_probs3 + eps)

                job_logits1 = np.log(job_act_probs1 + eps)
                job_logits2 = np.log(job_act_probs2 + eps)
                job_logits3 = np.log(job_act_probs3 + eps)

                combined_node_logits = w1 * node_logits1 + w2 * node_logits2 + w3 * node_logits3
                combined_job_logits = w1 * job_logits1 + w2 * job_logits2 + w3 * job_logits3

                # apply mask and Gumbel sampling on combined logits
                node_act_logits = combined_node_logits + (node_valid_mask - 1) * 10000.0
                node_act = np.argmax(node_act_logits, axis=1)

                job_valid_mask_re = job_valid_mask.reshape(1, np.shape(job_act_probs1)[1], np.shape(job_act_probs1)[2])
                job_act_logits = combined_job_logits + (job_valid_mask_re - 1) * 10000.0
                job_act = np.argmax(job_act_logits, axis=2)

                node = node_action_map[node_act[0]]
                job_idx = job_dags.index(node.job_dag)
                mec = mec_action_map[job_act[0, job_idx]]

                obs, reward1, reward2, reward3, done = env.step2(node, mec)
                # 在每次决策后打印关键信息
                # print(f"时间步 {env.wall_time}:")
                # print(f"  可用节点: {len(free_mecs)}, 可用任务: {len(frontier_nodes)}")
                # print(f"  选择节点 {node}")
                # print(f"  选择节点 {node}, 任务 {job_idx}, MEC {mec}")
                # print(f"  节点概率1: {node_act_probs1}, 节点概率2: {node_act_probs2},节点概率3: {node_act_probs3},最终节点概率: {node_act_probs}")
                # print(f"  mec概率: {job_act_probs}")

            # 计算指标
            d = np.sum([j.completion_time - j.start_time for j in env.finished_job_dags])
            h = len(env.failed_job_dags)
            s = np.sum([max(0, j.completion_time - (j.ddl + j.start_time))
                        for j in env.finished_job_dags if j.var == "soft"])
            total = d + h + s
            self.save(w1, w2, w3, total, d, h, s)

            totals.append(total)
            ds.append(d)
            hs.append(h)
            ss.append(s)
            i += 1

        # 计算平均
        avg_total = np.mean(totals)
        avg_d = np.mean(ds)
        avg_h = np.mean(hs)
        avg_s = np.mean(ss)
        # 保存到 excel
        self.save(w[0], w[1], w[2], avg_total, avg_d, avg_h, avg_s)

        # 返回结果给训练过程使用（可选）
        return avg_total, avg_d, avg_h, avg_s
